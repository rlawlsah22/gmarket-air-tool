"""
에어프레미아 항공료 체커
========================

에어프레미아(airpremia.com) 공식 홈페이지에서 특정 출발일 기준으로
- 목적지 / 출발일 / 박수(귀국일) / 인원수 를 입력하면
- 실제 예약 화면과 동일한 흐름(목적지→날짜→인원 검색 → 출발편 이코노미
  선택 → 위탁수하물 23kg 1개 포함 요금제 중 최저가 선택 → 귀국편도 동일 →
  총 결제금액 확인)을 그대로 따라가서
- "1인당 항공료(만원 단위 올림)" 를 계산해주는 도구입니다.

전제 조건
---------
1. Python 3.9+
2. pip install selenium openpyxl
3. Chrome + 해당 버전에 맞는 chromedriver (Selenium 4.6+ 는 Selenium Manager가
   자동으로 chromedriver를 받아오므로 별도 설치 없이도 대부분 동작합니다.)

사용 방법 (커맨드라인 예시)
--------------------------
    python airpremia_fare_checker.py --dest BKK --start 2026-08-04 --nights 3 --adt 4

여러 날짜를 한 번에 훑으려면 --start-range / --end-range 를 사용하세요.
    python airpremia_fare_checker.py --dest BKK --start-range 2026-08-01 \
        --end-range 2026-08-31 --nights 3 --adt 4 --out fares.xlsx

동작 원리 메모 (사이트 리버스엔지니어링 결과)
--------------------------------------------
- 검색 결과 페이지는 아래 URL 패턴으로 바로 진입 가능합니다 (SPA지만 새로고침 시
  쿼리 파라미터 기준으로 다시 렌더링됨을 확인했습니다):

    https://www.airpremia.com/a/ko/ticket/flight/departure
        ?ADT={인원}&CHD=0&INF=0&INS=0
        &START_DATE={출발일 YYYY-MM-DD}&END_DATE={귀국일 YYYY-MM-DD}
        &ORG=ICN&DES={목적지코드}
        &CURRENCY=KRW&TRIP_TYPE=RT&IS_POINT=false&LANG=ko
        &SEAT_CLASS_TYPE=&AGENT=&EVENT_PROMOTION_CODE=

  귀국편 선택 화면은 동일 파라미터의 .../flight/return 경로입니다.

- 목적지에 따라(예: 태국) "태국으로 입국하시는 손님께 안내 드립니다" 라는
  안내 팝업이 뜨며, 체크박스를 클릭해 "확인" 버튼이 활성화된 후 눌러야
  다음 진행이 가능합니다. 이 스크립트는 팝업이 있으면 자동으로 처리하고,
  없으면 그냥 넘어갑니다.

- 이코노미/와이드프리미엄 클릭 시 각 등급 밑에 "스탠다드" / "플렉스" 같은
  요금제 카드가 펼쳐지며, 각 카드에 위탁수하물 정보가 표기됩니다.
  이 스크립트는 "23" 과 "1개" 문구가 모두 포함된(=23kg 1개 포함) 요금제만
  후보로 두고, 그중 가장 저렴한 걸 자동으로 선택합니다.

- 특정 날짜에 운항 편이 없으면 "운항하지 않는 날이에요" 메시지가 뜹니다.
  이 경우 해당 날짜는 결과에 '운항없음'으로 표시하고 다음 날짜로 넘어갑니다.

- API 참고용 (직접 호출은 Referer 등 헤더 문제로 404가 날 수 있어
  이 스크립트는 브라우저 자동화 방식을 기본으로 합니다):
    GET /api/v1/low-fares?origin=&destination=&beginDate=&endDate=
        &adtCount=&chdCount=&infCount=&insCount=&tripType=RT&usePoint=false&useCache=true
    GET /api/v1/fares?origin=&destination=&departureDate=&returnDate=
        &usePoint=false&promotionCode=&adtCount=&chdCount=&infCount=&insCount=
"""

from __future__ import annotations

import argparse
import math
import re
import time
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Optional

from selenium import webdriver
# PyInstaller가 selenium.webdriver의 지연(lazy) import를 정적 분석으로 잡지 못해
# exe 빌드 시 'No module named selenium.webdriver.chrome.options' 같은 에러가 나는
# 문제가 있어, 실제 쓰는 서브모듈들을 아래에서 명시적으로 한 번 더 import 해둡니다.
import selenium.webdriver.chrome.options  # noqa: F401
import selenium.webdriver.chrome.service  # noqa: F401
import selenium.webdriver.chrome.webdriver  # noqa: F401
import selenium.webdriver.common.by  # noqa: F401
import selenium.webdriver.support.expected_conditions  # noqa: F401
import selenium.webdriver.support.ui  # noqa: F401
from selenium.common.exceptions import (
    ElementClickInterceptedException,
    NoSuchElementException,
    StaleElementReferenceException,
    TimeoutException,
)
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

BASE_URL = "https://www.airpremia.com"
ORIGIN = "ICN"  # 출발지는 서울(인천) 고정


@dataclass
class FareResult:
    departure_date: str
    return_date: str
    status: str  # "OK" | "NO_FLIGHT" | "ERROR"
    total_amount: Optional[int] = None
    per_person: Optional[int] = None
    per_person_rounded: Optional[int] = None
    note: str = ""


def build_leg_url(leg: str, dest: str, adt: int, start_date: str, end_date: str) -> str:
    """leg: 'departure' or 'return'"""
    return (
        f"{BASE_URL}/a/ko/ticket/flight/{leg}"
        f"?ADT={adt}&CHD=0&INF=0&INS=0"
        f"&START_DATE={start_date}&END_DATE={end_date}"
        f"&ORG={ORIGIN}&DES={dest}"
        f"&CURRENCY=KRW&TRIP_TYPE=RT&IS_POINT=false&LANG=ko"
        f"&SEAT_CLASS_TYPE=&AGENT=&EVENT_PROMOTION_CODE="
    )


def round_up_to_10000(amount: int) -> int:
    return math.ceil(amount / 10000) * 10000


def parse_won(text: str) -> Optional[int]:
    """'KRW 3,181,200' / '3,181,200원' 등에서 숫자만 뽑아 int로."""
    digits = re.sub(r"[^\d]", "", text)
    return int(digits) if digits else None


def dismiss_thailand_notice(driver, wait_sec: float = 4.0) -> None:
    """목적지에 따라 뜨는 입국 안내 팝업을 처리. 없으면 그냥 넘어감."""
    try:
        checkbox = WebDriverWait(driver, wait_sec).until(
            EC.presence_of_element_located(
                (By.XPATH, "//*[contains(text(), '위 내용을 확인하였습니다')]")
            )
        )
        # 라벨/텍스트 근처의 실제 체크박스 input을 찾는다
        try:
            cb_input = driver.find_element(
                By.XPATH,
                "//*[contains(text(), '위 내용을 확인하였습니다')]"
                "/preceding::input[@type='checkbox'][1]",
            )
        except NoSuchElementException:
            cb_input = checkbox  # 최후 수단으로 텍스트 요소 자체 클릭 시도

        driver.execute_script("arguments[0].click();", cb_input)

        confirm_btn = WebDriverWait(driver, wait_sec).until(
            EC.element_to_be_clickable(
                (By.XPATH, "//button[contains(., '확인')]")
            )
        )
        confirm_btn.click()
        time.sleep(0.5)
    except TimeoutException:
        # 팝업이 없는 목적지면 정상
        pass


def _find_date_candidates(driver, label: str):
    elems = driver.find_elements(
        By.XPATH, f"//*[contains(normalize-space(string(.)), '{label}')]"
    )
    return [e for e in elems if e.text.strip().startswith(label)]


def _visible_calendar_arrows(driver):
    """
    좌우 화살표(svg.cursor-pointer)를 찾아 x좌표 순으로 [이전, 다음] 순서로 반환.
    실제 사이트 마크업 확인 결과, 날짜 캐러셀의 좌우 화살표는 <svg class="... cursor-pointer">
    형태이고 텍스트가 없어 By.XPATH 텍스트 검색으로는 못 찾는다.
    """
    svgs = driver.find_elements(By.CSS_SELECTOR, "svg.cursor-pointer")
    visible = []
    for s in svgs:
        try:
            if s.is_displayed():
                visible.append(s)
        except StaleElementReferenceException:
            continue
    visible.sort(key=lambda e: e.location.get("x", 0))
    return visible


def select_date_tile(driver, target: date, wait_sec: float = 12.0, max_page_clicks: int = 24) -> None:
    """
    URL 파라미터로 날짜를 넣어도 사이트가 자동으로 항공편 카드를 펼쳐주지 않는
    경우가 있어(캘린더만 뜨고 '이코노미' 버튼이 안 나타남), 캘린더에서 해당
    날짜 타일을 실제로 한 번 클릭해줘야 한다.

    캘린더는 한 번에 7일 정도만 보여주고, 원하는 날짜가 그 창 밖에 있으면
    좌우 화살표(<svg class="... cursor-pointer">)를 눌러 페이지를 넘겨야
    나타난다 (직접 화면 확인 및 라이브 테스트로 확인함 — 연속으로 여러 날짜를
    조회하다 보면 캘린더가 이전 조회의 위치에 걸쳐 있을 때가 있어, 매번 목표
    날짜가 첫 화면에 보인다는 보장이 없다).
    """
    label = f"{target.month}.{target.day}"  # 예: '8.4'

    def _pick_best(candidates):
        articles = [e for e in candidates if e.tag_name.lower() == "article"]
        visible_articles = []
        for e in articles:
            try:
                if e.is_displayed():
                    visible_articles.append(e)
            except StaleElementReferenceException:
                continue
        if visible_articles:
            return visible_articles[0]
        if articles:
            return articles[0]
        if candidates:
            candidates.sort(key=lambda e: len(e.text or ""))
            return candidates[0]
        return None

    # 캘린더 자체가 뜰 때까지는 기다린다 (article이 최소 1개는 있어야 함)
    try:
        WebDriverWait(driver, wait_sec).until(
            lambda d: d.find_elements(By.TAG_NAME, "article")
        )
    except TimeoutException:
        raise

    tile = _pick_best(_find_date_candidates(driver, label))

    if tile is None:
        # 목표 날짜가 현재 보이는 창 밖에 있음 → 방향을 추정해서 화살표로 페이지 이동
        sample_articles = [
            e for e in driver.find_elements(By.TAG_NAME, "article") if e.is_displayed()
        ]
        direction = "next"
        if sample_articles:
            m = re.match(r"(\d+)\.(\d+)", sample_articles[0].text.strip())
            if m:
                samp = (int(m.group(1)), int(m.group(2)))
                tgt = (target.month, target.day)
                # 12월→1월 연말 경계는 다루지 않음 (한 번에 몇 주 이상 넘기는 상황은 없다고 가정)
                direction = "next" if tgt > samp else "prev"

        arrows = _visible_calendar_arrows(driver)
        if len(arrows) < 2:
            raise TimeoutException(
                f"{label} 날짜 타일을 찾지 못했고, 페이지 이동 화살표도 찾지 못함"
            )

        for _ in range(max_page_clicks):
            arrows = _visible_calendar_arrows(driver)
            if len(arrows) < 2:
                break
            arrow = arrows[1] if direction == "next" else arrows[0]
            try:
                arrow.click()
            except (ElementClickInterceptedException, StaleElementReferenceException):
                driver.execute_script(
                    "arguments[0].dispatchEvent(new MouseEvent('click', {bubbles:true}));",
                    arrow,
                )
            time.sleep(0.6)
            tile = _pick_best(_find_date_candidates(driver, label))
            if tile is not None:
                break

    if tile is None:
        raise TimeoutException(f"{label} 날짜 타일을 페이지 이동으로도 찾지 못함")

    try:
        tile.click()
    except (ElementClickInterceptedException, StaleElementReferenceException):
        driver.execute_script("arguments[0].click();", tile)
    time.sleep(0.8)


def wait_for_flight_state(driver, wait_sec: float = 12.0) -> str:
    """
    'OK' | 'NO_FLIGHT' 반환.

    주의: "운항하지 않는 날이에요" 라는 문구는 Next.js가 미리 심어둔 hydration용
    <script> 태그(JSON payload) 안에도 항상 포함돼 있어서, 단순히
    contains(text(), '운항하지 않는 날') 로만 찾으면 실제 그 날짜가 운항하는지와
    무관하게 '항상' 걸린다 (실제로 라이브 테스트에서 확인함). 그래서:
      1) 이코노미 버튼이 있으면 무조건 OK로 판단 (최우선)
      2) 이코노미 버튼이 없을 때만, '화면에 보이는(is_displayed)' 요소 중에
         해당 문구가 있는지로 NO_FLIGHT 여부를 판단한다.
    """
    try:
        WebDriverWait(driver, wait_sec).until(
            lambda d: d.find_elements(By.XPATH, "//button[contains(., '이코노미')]")
            or d.find_elements(
                By.XPATH,
                "//*[contains(text(), '운항하지 않는 날') and not(self::script) and not(self::style)]",
            )
        )
    except TimeoutException:
        return "ERROR"

    if driver.find_elements(By.XPATH, "//button[contains(., '이코노미')]"):
        return "OK"

    no_fly_candidates = driver.find_elements(
        By.XPATH,
        "//*[contains(text(), '운항하지 않는 날') and not(self::script) and not(self::style)]",
    )
    for el in no_fly_candidates:
        try:
            if el.is_displayed():
                return "NO_FLIGHT"
        except StaleElementReferenceException:
            continue
    return "ERROR"


def select_cheapest_valid_fare(driver, wait_sec: float = 10.0) -> None:
    """
    '이코노미' 탭을 열고, 위탁수하물 23kg 1개가 포함된 요금제(보통 스탠다드/
    플렉스) 카드들 중 가장 저렴한 걸 골라 '선택하기' 버튼을 클릭한다.
    """
    economy_btn = WebDriverWait(driver, wait_sec).until(
        EC.element_to_be_clickable((By.XPATH, "//button[contains(., '이코노미')]"))
    )
    economy_btn.click()
    time.sleep(0.8)  # 요금제 카드 펼쳐지는 애니메이션 대기

    # '선택하기' 버튼들을 요금제 카드 단위로 순회
    select_buttons = WebDriverWait(driver, wait_sec).until(
        EC.presence_of_all_elements_located(
            (By.XPATH, "//button[contains(., '선택하기')]")
        )
    )

    candidates = []  # (price:int, button)
    for btn in select_buttons:
        try:
            # 버튼을 감싸는 요금제 카드(공통 조상)에서 가격/수하물 텍스트를 찾는다
            card = btn.find_element(
                By.XPATH,
                "./ancestor::*[.//*[contains(text(), '위탁수하물')]][1]",
            )
        except NoSuchElementException:
            continue

        card_text = card.text
        if "위탁수하물" not in card_text:
            continue
        if not (re.search(r"23\s*kg", card_text, re.IGNORECASE) and "1개" in card_text):
            continue  # 23kg 1개 포함 요금제가 아니면 후보에서 제외

        price_match = re.search(r"[\d,]{5,}", card_text)
        if not price_match:
            continue
        price = parse_won(price_match.group())
        if price:
            candidates.append((price, btn))

    if not candidates:
        raise RuntimeError("23kg 1개 포함 요금제를 찾지 못했습니다 (페이지 구조 변경 가능성)")

    candidates.sort(key=lambda x: x[0])
    cheapest_btn = candidates[0][1]
    try:
        cheapest_btn.click()
    except ElementClickInterceptedException:
        driver.execute_script("arguments[0].click();", cheapest_btn)
    time.sleep(0.5)


def click_next_leg(driver, wait_sec: float = 10.0) -> None:
    """출발편 선택 후 '다음 여정' 버튼 클릭 (귀국편 화면으로 이동)."""
    next_btn = WebDriverWait(driver, wait_sec).until(
        EC.element_to_be_clickable((By.XPATH, "//button[contains(., '다음 여정')]"))
    )
    next_btn.click()
    time.sleep(1.0)


def read_total_amount(driver, wait_sec: float = 10.0) -> int:
    """'총 결제금액' 옆의 금액을 읽어온다."""
    el = WebDriverWait(driver, wait_sec).until(
        EC.presence_of_element_located(
            (
                By.XPATH,
                "//*[contains(text(), '총 결제금액')]"
                "/following::*[contains(text(), 'KRW') or contains(text(), '원')][1]",
            )
        )
    )
    amount = parse_won(el.text)
    if amount is None:
        raise RuntimeError(f"총 결제금액 파싱 실패: {el.text!r}")
    return amount


def _select_date_tile_with_retry(driver, target: date, attempts: int = 3) -> None:
    """
    select_date_tile()이 실패하면(캘린더가 예상보다 느리게 뜨는 등의 일시적인
    이유일 수 있어) 페이지를 새로고침하고 태국 안내 팝업을 다시 처리한 뒤
    재시도한다. 마지막 시도까지 실패하면 TimeoutException을 그대로 던진다.
    """
    last_err = None
    for i in range(attempts):
        try:
            select_date_tile(driver, target)
            return
        except TimeoutException as e:
            last_err = e
            if i < attempts - 1:
                driver.refresh()
                time.sleep(1.5)
                dismiss_thailand_notice(driver)
    raise last_err


def check_one_date(
    driver,
    dest: str,
    departure_date: date,
    nights: int,
    adt: int,
) -> FareResult:
    return_date = departure_date + timedelta(days=nights)
    dep_str = departure_date.strftime("%Y-%m-%d")
    ret_str = return_date.strftime("%Y-%m-%d")

    try:
        # --- 출발편 ---
        driver.get(build_leg_url("departure", dest, adt, dep_str, ret_str))
        dismiss_thailand_notice(driver)

        try:
            _select_date_tile_with_retry(driver, departure_date)
        except TimeoutException:
            return FareResult(
                dep_str, ret_str, "ERROR", note="출발일 날짜 타일을 찾지 못함(캘린더 범위 밖일 수 있음)"
            )

        state = wait_for_flight_state(driver)
        if state == "NO_FLIGHT":
            return FareResult(dep_str, ret_str, "NO_FLIGHT", note="출발편 운항 없음")
        if state == "ERROR":
            return FareResult(dep_str, ret_str, "ERROR", note="출발편 페이지 로딩 실패")

        select_cheapest_valid_fare(driver)
        click_next_leg(driver)

        # --- 귀국편 (다음 여정 클릭 시 보통 자동 이동되지만, 안전하게 URL도 확인) ---
        dismiss_thailand_notice(driver)

        try:
            _select_date_tile_with_retry(driver, return_date)
        except TimeoutException:
            return FareResult(
                dep_str, ret_str, "ERROR", note="귀국일 날짜 타일을 찾지 못함(캘린더 범위 밖일 수 있음)"
            )

        state = wait_for_flight_state(driver)
        if state == "NO_FLIGHT":
            return FareResult(dep_str, ret_str, "NO_FLIGHT", note="귀국편 운항 없음")
        if state == "ERROR":
            return FareResult(dep_str, ret_str, "ERROR", note="귀국편 페이지 로딩 실패")

        select_cheapest_valid_fare(driver)

        total = read_total_amount(driver)
        per_person = total / adt
        per_person_rounded = round_up_to_10000(int(math.ceil(per_person)))

        return FareResult(
            dep_str,
            ret_str,
            "OK",
            total_amount=total,
            per_person=int(round(per_person)),
            per_person_rounded=per_person_rounded,
        )

    except (TimeoutException, StaleElementReferenceException, RuntimeError) as e:
        return FareResult(dep_str, ret_str, "ERROR", note=str(e))


def init_driver(headless: bool = False):
    """G마켓 도구와 동일한 패턴: driver 생성만 담당. GUI 등 상위 코드가 lifecycle을 관리."""
    options = webdriver.ChromeOptions()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--window-size=1400,1000")
    return webdriver.Chrome(options=options)


def run(
    dest: str,
    nights: int,
    adt: int,
    start_date: date,
    end_date: Optional[date] = None,
    headless: bool = False,
) -> list[FareResult]:
    """
    start_date 하나만 넘기면 그 날짜 하나만 체크.
    end_date 까지 넘기면 start_date~end_date 를 하루 단위로 순회하며 체크.
    (CLI 용. GUI에서는 init_driver()로 직접 드라이버를 관리하고 check_one_date()를
    반복 호출하는 방식을 쓴다 — 여러 세트를 하나의 브라우저 세션으로 처리하기 위해.)
    """
    driver = init_driver(headless=headless)

    results: list[FareResult] = []
    try:
        current = start_date
        last = end_date or start_date
        while current <= last:
            result = check_one_date(driver, dest, current, nights, adt)
            results.append(result)
            print(
                f"[{result.departure_date} -> {result.return_date}] "
                f"{result.status} "
                + (
                    f"총액={result.total_amount:,} / 1인={result.per_person_rounded:,}"
                    if result.status == "OK"
                    else result.note
                )
            )
            current += timedelta(days=1)
    finally:
        driver.quit()

    return results


def write_styled_sheet(ws, rows: list[FareResult], adt: int) -> None:
    """
    G마켓항공료추출기 엑셀 파일과 동일한 스타일로 시트를 채운다.
    - 헤더: 맑은 고딕 10 굵게, 흰 글씨, 남색(#1F4E79) 배경, 가운데 정렬, 얇은 테두리
    - 데이터: 연한 초록(#E2EFDA) 배경, 가운데 정렬, 얇은 테두리, 금액은 #,##0 서식
    - 비고: 운항없음/에러인 경우 그 사유를 표시, 정상 조회된 날짜는 비워둠
    - 1인금액: 총금액 ÷ 인원수를 만원 단위로 올림한 최종 값 (raw 금액은
      FareResult.per_person 에서 코드로만 참조 가능, 엑셀엔 표시하지 않음)
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    DATA_FILL = PatternFill("solid", fgColor="E2EFDA")
    CENTER = Alignment(horizontal="center", vertical="center")
    THIN = Side(style="thin")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    headers = ["출발일", "귀국일", "상태", f"{adt}인총금액", "1인금액", "비고"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    for r in rows:
        note = "" if r.status == "OK" else r.note
        ws.append([
            r.departure_date,
            r.return_date,
            r.status,
            r.total_amount if r.total_amount is not None else "",
            r.per_person_rounded if r.per_person_rounded is not None else "",
            note,
        ])

    money_cols = (4, 5)  # D~E: 총금액 / 1인금액
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = DATA_FILL
            cell.alignment = CENTER
            cell.border = BORDER
            if col_idx in money_cols and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    widths = {"A": 13, "B": 13, "C": 10, "D": 16, "E": 12, "F": 20}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def export_to_excel(results: list[FareResult], out_path: str, adt: int = 1) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "항공료체크"
    write_styled_sheet(ws, results, adt)
    wb.save(out_path)
    print(f"엑셀 저장 완료: {out_path}")


def main():
    parser = argparse.ArgumentParser(description="에어프레미아 항공료 체커")
    parser.add_argument("--dest", required=True, help="목적지 공항코드 (예: BKK)")
    parser.add_argument("--nights", type=int, required=True, help="박수 (예: 3)")
    parser.add_argument("--adt", type=int, required=True, help="성인 인원수")
    parser.add_argument("--start", help="단일 출발일 YYYY-MM-DD")
    parser.add_argument("--start-range", help="범위 조회 시작 출발일 YYYY-MM-DD")
    parser.add_argument("--end-range", help="범위 조회 종료 출발일 YYYY-MM-DD")
    parser.add_argument("--out", default="airpremia_fares.xlsx", help="엑셀 출력 파일명")
    parser.add_argument("--headless", action="store_true", help="브라우저 창 없이 실행")
    args = parser.parse_args()

    if args.start:
        start_date = datetime.strptime(args.start, "%Y-%m-%d").date()
        end_date = None
    elif args.start_range and args.end_range:
        start_date = datetime.strptime(args.start_range, "%Y-%m-%d").date()
        end_date = datetime.strptime(args.end_range, "%Y-%m-%d").date()
    else:
        parser.error("--start 또는 (--start-range 와 --end-range) 를 지정하세요.")
        return

    results = run(
        dest=args.dest,
        nights=args.nights,
        adt=args.adt,
        start_date=start_date,
        end_date=end_date,
        headless=args.headless,
    )
    export_to_excel(results, args.out, adt=args.adt)


if __name__ == "__main__":
    main()
