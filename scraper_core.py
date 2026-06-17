"""
scraper_core.py
G마켓 항공료 자동 추출 - 핵심 스크래핑 엔진
"""

import time
import re
from datetime import date, timedelta
from typing import Optional

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  상수
# ─────────────────────────────────────────────
AIRPORTS = {
    "국내": {
        "인천": "ICN", "부산": "PUS", "김포": "GMP",
        "청주": "CJJ", "대구": "TAE",
    },
    "일본": {
        "삿포로": "CTS", "아오모리": "AOJ", "센다이": "SDJ",
        "도쿄(나리타)": "NRT", "도쿄(하네다)": "HND", "시즈오카": "FSZ", "나고야": "NGO",
        "고마츠": "KMQ", "오사카": "OSA", "고베": "UKB",
        "히로시마": "HIJ", "다카마츠": "TAK", "마쓰야마": "MYJ",
        "후쿠오카": "FUK", "기타큐슈": "KKJ", "사가": "HSG",
        "나가사키": "NGS", "구마모토": "KMJ", "미야자키": "KMI",
        "가고시마": "KOJ", "오키나와": "OKA",
    },
    "중국": {
        "북경": "BJS", "천진": "TSN", "청도": "TAO",
        "연태": "YNT", "위해": "WEH", "대련": "DLC",
        "심양": "SHE", "제남": "TNA", "염성": "YNZ",
        "정주": "CGO", "하이커우": "HAK", "싼야": "SYX",
        "곤명": "KMG", "광저우": "CAN", "심펀": "SZX", "하문": "XMN",
    },
    "태국": {
        "방콕": "BKK", "치앙마이": "CNX", "푸켓": "HKT",
    },
    "베트남": {
        "하노이": "HAN", "다낭": "DAD", "나트랑": "CXR",
        "호치민": "SGN", "푸꾸옥": "PQC",
    },
    "필리핀": {
        "마닐라": "MNL", "클락": "CRK", "세부": "CEB", "보라카이": "MPH",
    },
}

# 도시명→코드 플랫 딕셔너리 (역방향 포함)
AIRPORT_CODES = {}
for country, cities in AIRPORTS.items():
    for city, code in cities.items():
        AIRPORT_CODES[city] = code
        AIRPORT_CODES[code] = code

LCC_TIER1 = ["진에어", "이스타항공", "티웨이항공"]
LCC_TIER2 = ["제주항공", "에어부산"]
LCC_OTHER = ["에어로케이", "에어프레미아", "에어서울", "파라타항공"]
LCC_ALL   = LCC_TIER1 + LCC_TIER2 + LCC_OTHER
FSC_ALL   = ["아시아나항공", "대한항공"]
FOREIGN_ALL = [
    "JAL", "ANA", "피치항공", "제트스타재팬", "스프링재팬",
    "중국국제항공", "중국남방항공", "중국동방항공", "산동항공", "샤먼항공", "하이난항공",
    "베트남항공", "비엣젯항공", "뱀부항공",
    "타이항공", "타이에어아시아", "방콕에어웨이즈",
    "에어아시아", "세부퍼시픽", "필리핀항공",
]

TARGET_SELLERS = ["롯데관광", "롯데제이티비", "여행이지"]

TIME_BANDS = {
    "새벽": (0,  6),
    "오전": (6,  12),
    "오후": (12, 18),
    "야간": (18, 24),
    "전체": (0,  24),
}

JS_PARSE = """
const TARGET_SELLERS = ['롯데관광', '롯데제이티비', '여행이지'];
const results = [];
const cards = document.querySelectorAll('.box__item-card');
cards.forEach((card, idx) => {
    const airlineEls = card.querySelectorAll('.text__airline');
    const airline = airlineEls.length > 0 ? airlineEls[0].innerText.trim() : '';
    const timeEls = card.querySelectorAll('.box__time-info .text__time');
    const dep  = timeEls[0] ? timeEls[0].innerText.trim() : '';
    const arr  = timeEls[1] ? timeEls[1].innerText.trim() : '';
    const rDep = timeEls[2] ? timeEls[2].innerText.trim() : '';
    const rArr = timeEls[3] ? timeEls[3].innerText.trim() : '';

    // 오는편 비행시간 파싱
    const summaryEls = card.querySelectorAll('.box__summary-info .text__time');
    let rDuration = 0;
    if (summaryEls.length >= 2) {
        const txt = summaryEls[1].innerText.trim();
        const hm = txt.match(/(\\d+)시간\\s*(\\d*)분?/);
        if (hm) rDuration = parseInt(hm[1]) * 60 + (hm[2] ? parseInt(hm[2]) : 0);
    }

    // 오는편 날짜 파싱 (link의 key 파라미터에서 추출)
    let rDepDate = '';
    const link = card.querySelector('a.link__seller-select');
    if (link) {
        const key = link.href.match(/key=([^&]+)/);
        if (key) {
            const parts = key[1].split('-');
            if (parts.length >= 2) {
                const d = parts[1].substring(0, 8);
                rDepDate = d.substring(0,4) + '-' + d.substring(4,6) + '-' + d.substring(6,8);
            }
        }
    }

    // 카드 상단 표시가
    const cardPriceEl = card.querySelector('.box__discount-cost .text__price') ||
                        card.querySelector('.box__seller-cost .text__price');
    const cardPrice = cardPriceEl ? cardPriceEl.innerText.trim() : '';

    if (airline && dep && arr) {
        results.push({airline, dep, arr, rDep, rArr, rDepDate, rDuration, cardPrice});
    }
});
return results;
"""

# ─────────────────────────────────────────────
#  디버그: 카드 구조 진단용 JS
# ─────────────────────────────────────────────
JS_DEBUG_CARD = """
const cards = document.querySelectorAll('.box__item-card');
if (!cards.length) return JSON.stringify({error: '카드 없음', url: location.href});

const c = cards[0];

// .text__price 셀렉터 후보들 전부 시도
const selectors = [
    '.box__discount-cost .text__price',
    '.box__seller-cost .text__price',
    '.text__price',
    '.box__cost .text__price',
    '.box__total-cost .text__price',
    '[class*="price"]',
    '[class*="cost"]',
];
const selectorResults = {};
selectors.forEach(sel => {
    try {
        const el = c.querySelector(sel);
        selectorResults[sel] = el ? el.innerText.trim() : null;
    } catch(e) {
        selectorResults[sel] = 'ERROR: ' + e.message;
    }
});

// 카드 내 class에 price/cost 포함된 요소 전부 수집
const allPriceLike = [...c.querySelectorAll('*')]
    .filter(el => (el.className||'').toString().match(/price|cost/i) && el.children.length === 0)
    .slice(0, 10)
    .map(el => ({class: el.className, text: el.innerText.trim()}));

// 시간 요소
const times = [...c.querySelectorAll('.text__time')].map(e => e.innerText.trim()).slice(0,6);

// 항공사
const airline = c.querySelector('.text__airline')?.innerText.trim();

return JSON.stringify({
    카드수: cards.length,
    항공사: airline,
    시간들: times,
    셀렉터결과: selectorResults,
    가격유사요소: allPriceLike,
}, null, 2);
"""


# ─────────────────────────────────────────────
#  드라이버 초기화
# ─────────────────────────────────────────────
def init_driver(show: bool = False) -> webdriver.Chrome:
    opts = Options()
    if not show:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
    driver = webdriver.Chrome(options=opts)
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    })
    return driver


# ─────────────────────────────────────────────
#  URL 생성
# ─────────────────────────────────────────────
def build_url(origin: str, dest: str, dep_date: str, arr_date: str, adults: int = 4) -> str:
    base = "https://air.gmarket.co.kr/gm/init/srp/srpResultView.do"
    return (
        f"{base}?TTYPE=global&RTYPE=fromkr&SECTN=RT"
        f"&DSTAD={origin}&DEPCY=A"
        f"&ASTAD={dest}&ARRCY=A"
        f"&DDATE={dep_date}&ADATE={arr_date}"
        f"&NADT={adults}&NCHD=0&NINF=0"
        f"&CLS=Y&VIAYN=false&MSITE=P"
    )


# ─────────────────────────────────────────────
#  필터 클릭
# ─────────────────────────────────────────────
AIRLINE_CODE_MAP = {
    "대한항공":   "KE",
    "아시아나항공": "OZ",
    "진에어":     "LJ",
    "제주항공":   "7C",
    "티웨이항공":  "TW",
    "이스타항공":  "ZE",
    "에어부산":   "BX",
    "에어서울":   "RS",
    "에어로케이":  "RF",
    "에어프레미아": "YP",
    "파라타항공":  "FD",
    "타이항공":   "TG",
    "필리핀항공":  "PR",
    "베트남항공":  "VN",
    "중국국제항공": "CA",
    "중국남방항공": "CZ",
    "중국동방항공": "MU",
    "산동항공":    "SC",
}

def click_filters(driver, specific_airlines=None):
    # 직항 체크박스 클릭 (check_flight_01)
    try:
        chk = driver.find_element(By.CSS_SELECTOR, "input#check_flight_01")
        if not chk.is_selected():
            driver.execute_script("arguments[0].click();", chk)
            time.sleep(1.5)
    except Exception:
        pass

    # opr=공동운항제외, bag=무료수하물
    for code in ["opr", "bag"]:
        try:
            btn = driver.find_element(By.CSS_SELECTOR, f"button[data-code='{code}']")
            if btn.get_attribute("aria-selected") == "false":
                driver.execute_script("arguments[0].click();", btn)
                time.sleep(1.5)
        except Exception:
            pass

    if specific_airlines:
        try:
            all_chk = driver.find_element(By.CSS_SELECTOR, "input#check_airline_0")
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", all_chk)
            time.sleep(1)
            if all_chk.is_selected():
                driver.execute_script("arguments[0].click();", all_chk)
                time.sleep(1)
            for airline_name in specific_airlines:
                code = AIRLINE_CODE_MAP.get(airline_name)
                if not code:
                    continue
                try:
                    chk = driver.find_element(By.CSS_SELECTOR, f"input[data-code='{code}']")
                    if not chk.is_selected():
                        driver.execute_script("arguments[0].click();", chk)
                        time.sleep(0.7)
                except Exception:
                    pass
            time.sleep(2)
        except Exception:
            pass


# ─────────────────────────────────────────────
#  시간 파싱
# ─────────────────────────────────────────────
def parse_hour(t: str) -> int:
    m = re.match(r"(\d{1,2}):", t)
    return int(m.group(1)) if m else -1


def in_band(t: str, band) -> bool:
    if isinstance(band, dict):
        if band["type"] == "range":
            h = parse_hour(t)
            m = int(re.search(r":(\d{2})", t).group(1)) if re.search(r":(\d{2})", t) else 0
            total_min = h * 60 + m
            return band["from"] <= total_min <= band["to"]
        elif band["type"] == "slots":
            return any(in_band(t, s) for s in band["slots"])
    if isinstance(band, list):
        return any(in_band(t, b) for b in band)
    if band == "전체":
        return True
    lo, hi = TIME_BANDS[band]
    h = parse_hour(t)
    return lo <= h <= hi


# ─────────────────────────────────────────────
#  금액 파싱
# ─────────────────────────────────────────────
def parse_price(s: str) -> int:
    return int(re.sub(r"[^0-9]", "", s)) if s else 0


def calc_per_person(total4: int) -> int:
    import math
    per = total4 / 4
    rounded = math.ceil(per / 10000) * 10000
    return rounded + 40000


# ─────────────────────────────────────────────
#  항공편 선택 로직
# ─────────────────────────────────────────────
JS_SELLER_PARSE = """
const TARGET_SELLERS = ['롯데관광', '롯데제이티비', '여행이지'];
let price = '';
let sellerName = '';
let minPrice = Infinity;

const sellerItems = document.querySelectorAll('.box__payment-item');
sellerItems.forEach(item => {
    const nameEl = item.querySelector('.text__seller-info');
    if (!nameEl) return;
    const name = nameEl.innerText.trim();
    if (!TARGET_SELLERS.includes(name)) return;
    const dcEl   = item.querySelector('.box__discount-cost .text__price');
    const costEl = item.querySelector('.box__seller-cost .text__price');
    const priceEl = dcEl || costEl;
    if (!priceEl) return;
    const val = parseInt(priceEl.innerText.replace(/[^0-9]/g, ''), 10);
    if (val < minPrice) {
        minPrice = val;
        price = priceEl.innerText.trim();
        sellerName = name;
    }
});

if (!price) {
    sellerItems.forEach(item => {
        const nameEl = item.querySelector('.text__seller-info');
        if (!nameEl) return;
        const dcEl   = item.querySelector('.box__discount-cost .text__price');
        const costEl = item.querySelector('.box__seller-cost .text__price');
        const priceEl = dcEl || costEl;
        if (!priceEl) return;
        const val = parseInt(priceEl.innerText.replace(/[^0-9]/g, ''), 10);
        if (val < minPrice) {
            minPrice = val;
            price = priceEl.innerText.trim();
            sellerName = nameEl.innerText.trim();
        }
    });
}
return {price, sellerName};
"""


def fetch_seller_price(driver, airline, dep, arr, card_price, log_fn=None) -> tuple:
    main_window = driver.current_window_handle
    try:
        seller_url = driver.execute_script("""
            const tgtAirline = arguments[0];
            const tgtDep = arguments[1];
            const tgtArr = arguments[2];
            const cards = document.querySelectorAll('.box__item-card');
            for(const card of cards) {
                const al = card.querySelector('.text__airline');
                const times = card.querySelectorAll('.box__time-info .text__time');
                if(!al || times.length < 2) continue;
                if(al.innerText.trim() === tgtAirline &&
                   times[0].innerText.trim() === tgtDep &&
                   times[1].innerText.trim() === tgtArr) {
                    const payItems = card.querySelectorAll('.box__payment-option .box__full-item');
                    for(const item of payItems) {
                        const cn = item.querySelector('.text__card');
                        if(cn && cn.innerText.trim() === '모든 결제수단') {
                            const link = item.querySelector('.link__seller-select');
                            if(link) {
                                return 'https://air.gmarket.co.kr' + link.pathname + link.search;
                            }
                        }
                    }
                }
            }
            return '';
        """, airline, dep, arr)

        if not seller_url:
            if log_fn:
                log_fn(f"      ⚠ 모든결제수단 URL 못찾음, fallback")
            return card_price, "fallback"

        if log_fn:
            log_fn(f"      → URL: ...{seller_url[-45:]}")

        driver.execute_script("window.open('about:blank', '_blank');")
        time.sleep(0.3)
        driver.switch_to.window(driver.window_handles[-1])
        driver.get(seller_url)
        try:
            WebDriverWait(driver, 12).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".box__payment-item .text__seller-info"))
            )
            WebDriverWait(driver, 12).until(
                lambda d: d.execute_script(
                    "return document.querySelectorAll('.box__payment-item .box__discount-cost .text__price').length > 0"
                )
            )
            time.sleep(1.5)
            result = driver.execute_script(JS_SELLER_PARSE)
            price  = result.get("price", "") if result else ""
            seller = result.get("sellerName", "") if result else ""
            if log_fn:
                log_fn(f"      → 판매사: {seller} / {price}")
        except Exception:
            price, seller = "", ""
        driver.close()
        driver.switch_to.window(main_window)
        if price:
            return price, seller
        return card_price, "fallback"
    except Exception as e:
        if log_fn:
            log_fn(f"      ⚠ 판매사 가격 오류: {e}")
        try:
            for h in driver.window_handles:
                if h != main_window:
                    driver.switch_to.window(h)
                    driver.close()
            driver.switch_to.window(main_window)
        except Exception:
            pass
        return card_price, "fallback"


def select_best(flights: list, config: dict) -> Optional[dict]:
    mode = config.get("airline_mode", "LCC우선_FSC대체")
    dep_band     = config.get("dep_band",     "전체")
    arr_band     = config.get("arr_band",     "전체")
    ret_dep_band = config.get("ret_dep_band", "전체")
    ret_arr_band = config.get("ret_arr_band", "전체")

    def time_ok(f):
        rDep = f.get("rDep", "")
        rArr = f.get("rArr", "")
        return (
            in_band(f["dep"],  dep_band) and
            in_band(f["arr"],  arr_band) and
            (not rDep or in_band(rDep, ret_dep_band)) and
            (not rArr or in_band(rArr, ret_arr_band))
        )

    def best_from(pool):
        candidates = [f for f in pool if time_ok(f)]
        if not candidates:
            return None
        return min(candidates, key=lambda f: parse_price(f.get("cardPrice", f.get("price", "0"))))

    if mode == "외항사만":
        return best_from([f for f in flights if f["airline"] in FOREIGN_ALL])
    if mode == "특정항공사":
        specific = config.get("specific_airlines", [])
        return best_from([f for f in flights if f["airline"] in specific])
    if mode == "LCC만":
        return best_from([f for f in flights if f["airline"] in LCC_ALL])
    if mode == "FSC만":
        return best_from([f for f in flights if f["airline"] in FSC_ALL])
    if mode == "LCC우선_FSC대체":
        result = best_from([f for f in flights if f["airline"] in LCC_TIER1])
        if result:
            return result
        result = best_from([f for f in flights if f["airline"] in LCC_TIER2])
        if result:
            return result
        result = best_from([f for f in flights if f["airline"] in LCC_OTHER])
        if result:
            return result
        return best_from([f for f in flights if f["airline"] in FSC_ALL])
    return None


# ─────────────────────────────────────────────
#  페이지 로드 및 파싱
# ─────────────────────────────────────────────

# 디버그 로그: 첫 번째 검색에서만 카드 구조 상세 출력
_debug_done = False

def fetch_flights(driver, url: str, log_fn=None, specific_airlines=None, airline_mode=None) -> list:
    global _debug_done

    driver.get(url)
    try:
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".box__item-card"))
        )
        WebDriverWait(driver, 20).until(
            lambda d: d.execute_script(
                "return document.querySelectorAll('.box__item-card .text__time').length > 0"
            )
        )
    except Exception:
        if log_fn:
            log_fn("  ⚠ 결과 로드 실패 (항공편 없을 수 있음)")
        return []

    time.sleep(3)

    click_filters(driver, specific_airlines=specific_airlines, airline_mode=airline_mode)

    # 필터 클릭 후 카드가 사라졌다가 다시 로드될 때까지 대기
    try:
        # 카드가 일단 사라지길 기다림 (필터 적용 중)
        time.sleep(2)
        WebDriverWait(driver, 8).until_not(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".box__item-card"))
        )
    except Exception:
        pass  # 안 사라져도 계속 진행

    try:
        # 카드가 다시 나타날 때까지 대기
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".box__item-card"))
        )
        WebDriverWait(driver, 20).until(
            lambda d: d.execute_script(
                "return document.querySelectorAll('.box__item-card .text__time').length > 0"
            )
        )
        time.sleep(4)
    except Exception:
        pass



    # 스크롤 다운으로 추가 항공편 로드 (모든 카드 로드될 때까지)
    last_count = 0
    for _ in range(10):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
        count = driver.execute_script(
            "return document.querySelectorAll('.box__item-card').length"
        )
        if count == last_count:
            break
        last_count = count

    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(2)
    try:
        flights = driver.execute_script(JS_PARSE)

        return flights or []
    except Exception as e:
        if log_fn:
            log_fn(f"  ⚠ JS 파싱 오류: {e}")
        return []


def reset_debug():
    """새 검색 시작 시 디버그 플래그 초기화"""
    global _debug_done
    _debug_done = False


# ─────────────────────────────────────────────
#  월 전체 수집
# ─────────────────────────────────────────────
def collect_monthly(origin: str, dest: str, year: int, month: int,
                    nights: int, config: dict,
                    show_browser: bool = False,
                    log_fn=None, progress_fn=None) -> list:
    import calendar
    _, last_day = calendar.monthrange(year, month)
    rows = []
    reset_debug()
    driver = init_driver(show=show_browser)

    try:
        for day in range(1, last_day + 1):
            dep_date = date(year, month, day)
            arr_date = dep_date + timedelta(days=nights)
            dep_str = dep_date.strftime("%Y%m%d")
            arr_str = arr_date.strftime("%Y%m%d")

            url = build_url(origin, dest, dep_str, arr_str)
            if log_fn:
                log_fn(f"  {dep_date.strftime('%m/%d')} ({dep_date.strftime('%a')}) 검색 중...")

            specific_airlines = config.get("specific_airlines", []) if config.get("airline_mode") == "특정항공사" else None
            flights = fetch_flights(driver, url, log_fn, specific_airlines=specific_airlines)
            best = select_best(flights, config) if flights else None

            if best:
                total4 = parse_price(best["price"])
                per1   = calc_per_person(total4)
                rows.append({
                    "dep_date":  dep_date.strftime("%Y-%m-%d"),
                    "arr_date":  arr_date.strftime("%Y-%m-%d"),
                    "airline":   best["airline"],
                    "dep":       best["dep"],
                    "arr":       best["arr"],
                    "rDep":      best.get("rDep", ""),
                    "rArr":      best.get("rArr", ""),
                    "total4":    total4,
                    "per1":      per1,
                    "seller":    best.get("sellerName", ""),
                    "found":     True,
                })
            else:
                rows.append({
                    "dep_date":  dep_date.strftime("%Y-%m-%d"),
                    "arr_date":  arr_date.strftime("%Y-%m-%d"),
                    "airline":   "",
                    "dep": "", "arr": "", "rDep": "", "rArr": "",
                    "total4": 0, "per1": 0, "seller": "",
                    "found": False,
                })

            if progress_fn:
                progress_fn(day, last_day)

    finally:
        driver.quit()

    return rows


# ─────────────────────────────────────────────
#  엑셀 저장
# ─────────────────────────────────────────────
FILL_HEADER  = PatternFill("solid", fgColor="1F4E79")
FILL_LCC     = PatternFill("solid", fgColor="E2EFDA")
FILL_TIER2   = PatternFill("solid", fgColor="FFF2CC")
FILL_NONE    = PatternFill("solid", fgColor="FCE4D6")
FILL_WHITE   = PatternFill("solid", fgColor="FFFFFF")

FONT_HEADER  = Font(color="FFFFFF", bold=True, name="맑은 고딕", size=10)
FONT_NONE    = Font(color="FF0000", bold=True, name="맑은 고딕", size=10)
FONT_NORMAL  = Font(name="맑은 고딕", size=10)

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")

HEADERS = ["출발일", "귀국편출발일", "귀국일", "항공사", "출발시간", "도착시간",
           "귀국출발", "귀국도착", "4인총금액(카드할인가)", "1인금액", "비고"]
COL_WIDTHS = [13, 13, 13, 12, 10, 10, 10, 10, 22, 18, 16]


def save_excel(rows: list, origin: str, dest: str,
               year: int, month: int, out_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{origin}_{dest}_{year}{month:02d}"
    ws.row_dimensions[1].height = 22

    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill   = FILL_HEADER
        cell.font   = FONT_HEADER
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w

    for r, row in enumerate(rows, 2):
        ws.row_dimensions[r].height = 18
        if row["found"]:
            airline = row["airline"]
            seller  = row.get("seller", "")
            if "⚠" in seller:
                fill = PatternFill("solid", fgColor="FCE4D6")
            elif airline in LCC_TIER1:
                fill = FILL_LCC
            elif airline in LCC_TIER2 or airline in FSC_ALL:
                fill = FILL_TIER2
            else:
                fill = FILL_LCC
            # 귀국편출발일: rDep이 rArr보다 크면(자정넘어) arr_date 전날, 아니면 arr_date
            rdep = row.get("rDep", "")
            rarr = row.get("rArr", "")
            try:
                import datetime as _dt
                arr_d = _dt.datetime.strptime(row["arr_date"], "%Y-%m-%d").date()
                if rdep and rarr:
                    rdep_min = int(rdep.split(":")[0]) * 60 + int(rdep.split(":")[1])
                    rarr_min = int(rarr.split(":")[0]) * 60 + int(rarr.split(":")[1])
                    # rDep > rArr 이면 자정넘어 도착 → 귀국편출발일 = arr_date
                    # rDep < rArr 이면 당일도착 → 귀국편출발일 = arr_date
                    # rDep이 00:xx이고 rArr이 07:xx → arr_date가 한국도착일이므로 현지출발은 arr_date
                    # 단, rArr < rDep이면 익일도착 → 현지출발은 arr_date 전날
                    if rarr_min < rdep_min:
                        rdep_date = (arr_d - _dt.timedelta(days=1)).strftime("%Y-%m-%d")
                    else:
                        rdep_date = row["arr_date"]
                else:
                    rdep_date = row["arr_date"]
            except Exception:
                rdep_date = row["arr_date"]

            values = [
                row["dep_date"], rdep_date, row["arr_date"], airline,
                row["dep"], row["arr"], row["rDep"], row["rArr"],
                row["total4"], row["per1"], row["seller"],
            ]
            font = FONT_NORMAL
        else:
            fill = FILL_NONE
            values = [row["dep_date"], "", row["arr_date"],
                      "X", "", "", "", "", "", "", ""]
            font = FONT_NONE

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill      = fill
            cell.font      = font
            cell.alignment = CENTER
            cell.border    = BORDER

        if row["found"]:
            ws.cell(row=r, column=9).number_format = '#,##0'
            ws.cell(row=r, column=10).number_format = '#,##0'

    ws.freeze_panes = "A2"
    wb.save(out_path)
